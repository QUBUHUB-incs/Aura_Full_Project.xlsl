from openpyxl import Workbook, load_workbook
import os

def write_to_aura_sheet(date, metrics, patches):
    sheet_path = "brain/aura_sheet.xlsx"

    if not os.path.exists(sheet_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Aura_Brain"
        ws.append(["Date", "Accuracy", "Speed", "Stability", "Patches"])
        wb.save(sheet_path)

    wb = load_workbook(sheet_path)
    ws = wb["Aura_Brain"]

    ws.append([
        str(date),
        round(metrics["accuracy"], 3),
        round(metrics["speed"], 3),
        round(metrics["stability"], 3),
        ", ".join(patches)
    ])

    wb.save(sheet_path)
    print(f"🧠 Logged learning data to Aura Sheet for {date}.")
